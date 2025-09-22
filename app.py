from flask import Flask, render_template, jsonify, send_file, Response, request
from jira_api import get_sprints, get_issues_with_details, get_sprint_details, get_sprint_name, get_task_summary, get_projects, get_boards_for_project, get_sprints_for_board, URL
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime
import pytz
import json  # Agregar esta importación
import csv
import io

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/projects')
def api_projects():
    try:
        projects = get_projects()
        return jsonify(projects)
    except requests.exceptions.ConnectTimeout:
        return jsonify({'error': 'Timeout connecting to Jira'}), 504
    except requests.exceptions.ReadTimeout:
        return jsonify({'error': 'Jira API read timed out'}), 504
    except requests.exceptions.HTTPError as e:
        return jsonify({'error': f'Jira HTTP error: {e.response.status_code}'}), e.response.status_code
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/projects/<int:project_id>/boards')
def api_project_boards(project_id):
    try:
        boards = get_boards_for_project(project_id)
        return jsonify(boards)
    except requests.exceptions.ConnectTimeout:
        return jsonify({'error': 'Timeout connecting to Jira'}), 504
    except requests.exceptions.ReadTimeout:
        return jsonify({'error': 'Jira API read timed out'}), 504
    except requests.exceptions.HTTPError as e:
        return jsonify({'error': f'Jira HTTP error: {e.response.status_code}'}), e.response.status_code
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/boards/<int:board_id>/sprints')
def api_board_sprints(board_id):
    try:
        sprints = get_sprints_for_board(board_id)
        return jsonify(sprints)
    except requests.exceptions.ConnectTimeout:
        return jsonify({'error': 'Timeout connecting to Jira'}), 504
    except requests.exceptions.ReadTimeout:
        return jsonify({'error': 'Jira API read timed out'}), 504
    except requests.exceptions.HTTPError as e:
        return jsonify({'error': f'Jira HTTP error: {e.response.status_code}'}), e.response.status_code
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/sprints')
def api_sprints():
    sprints = get_sprints()
    return jsonify(sprints)

@app.route('/api/sprints/<int:sprint_id>/details')
def api_sprint_details(sprint_id):
    sprint_details = get_sprint_details(sprint_id)
    issues = get_issues_with_details(sprint_id)
    task_summary = get_task_summary(issues)
    return jsonify({'sprint': sprint_details, 'issues': issues, 'task_summary': task_summary})

@app.route('/api/sprints/<int:sprint_id>/issues/download')
def download_sprint_issues(sprint_id):
    issues = get_issues_with_details(sprint_id)
    sprint_name = get_sprint_name(sprint_id)
    wb = Workbook()

    # Hoja de Worklogs
    ws_worklogs = wb.active
    ws_worklogs.title = "Worklogs"
    ws_worklogs.append(["Task Key", "Summary", "Type", "User", "Time Spent (hours)", "Link"])

    for issue in issues:
        issue_key = issue['key']
        issue_summary = issue['fields']['summary']
        issue_type = issue['fields']['issuetype']['name']
        issue_link = f"{URL}/browse/{issue_key}"
        if issue['worklogs']:
            for worklog in issue['worklogs']:
                ws_worklogs.append([issue_key, issue_summary, issue_type, worklog['author']['displayName'], worklog['timeSpentHours'], issue_link])
        else:
            ws_worklogs.append([issue_key, issue_summary, issue_type, "", 0, issue_link])

    # Hoja de Issues
    ws_issues = wb.create_sheet(title="Issues")
    ws_issues.append(["Task Key", "Summary", "Status", "Story Points"])

    for issue in issues:
        issue_key = issue['key']
        issue_summary = issue['fields']['summary']
        issue_status = issue['fields']['status']['name']
        story_points = issue['fields'].get('customfield_10016', 0)
        ws_issues.append([issue_key, issue_summary, issue_status, story_points])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(file_stream, as_attachment=True, download_name=f"{sprint_name}_worklogs.xlsx")

@app.route('/api/sprints/<int:sprint_id>/worklogs/download')
def download_sprint_worklogs(sprint_id):
    issues = get_issues_with_details(sprint_id)
    sprint_name = get_sprint_name(sprint_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Worklogs"
    
    # Encabezados
    headers = ["Task Key", "Summary", "Issue Type", "User", "Time Spent (hours)", "Link", "Started"]
    ws.append(headers)
    
    # Datos
    for issue in issues:
        if issue['worklogs']:
            for worklog in issue['worklogs']:
                ws.append([
                    issue['key'],
                    issue['fields']['summary'],
                    issue['fields']['issuetype']['name'],
                    worklog['author']['displayName'],
                    worklog['timeSpentHours'],
                    f"{URL}/browse/{issue['key']}",
                    worklog['started']
                ])
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f"{sprint_name}_detailed_worklogs.xlsx"
    )

def analyze_story_points_vs_time(story_points, time_spent):
    if story_points is None or story_points == 0:
        return ""
    
    # Definir rangos de horas para cada punto
    ranges = {
        1: (0, 2),
        2: (2, 4),
        3: (4, 8),
        5: (8, 16),
        8: (16, 24)
    }
    
    expected_range = ranges.get(story_points)
    if expected_range is None:
        return ""
    
    if time_spent < expected_range[0]:
        return "Sobre-estimado"
    elif time_spent > expected_range[1]:
        return "Sub-estimado"
    else:
        return "Correcto"

@app.route('/api/sprints/<int:sprint_id>/analysis/download')
def download_sprint_analysis(sprint_id):
    issues = get_issues_with_details(sprint_id)
    sprint_details = get_sprint_details(sprint_id)
    sprint_name = get_sprint_name(sprint_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sprint Analysis"
    
    headers = [
        "Issue Type",
        "Issue Key",
        "Summary",
        "Assignee",
        "Status",
        "Time Spent",
        "Story Points",
        "Story Points vs Time",
        "Parent Summary",
        "Fecha de Creación"
    ]
    ws.append(headers)
    
    # Obtener fecha de fin del sprint como datetime
    sprint_end = sprint_details.get('endDate')
    if sprint_end:
        # Ajustar la hora de cierre del sprint a las 23:59
        sprint_end_date = datetime.strptime(sprint_end, '%Y-%m-%d %H:%M:%S').date()
        sprint_end_dt = datetime.combine(sprint_end_date, datetime.max.time().replace(hour=23, minute=59, second=59, microsecond=0))
    else:
        sprint_end_dt = None
    
    for issue in issues:
        issue_type = issue['fields']['issuetype']['name']
        issue_key = issue['key']
        summary = issue['fields'].get('summary', '')
        
        assignee = ''
        if issue['fields'].get('assignee'):
            assignee = issue['fields']['assignee'].get('displayName', '')
        
        # --- Estado al final del sprint usando changelog ---
        status = issue['fields']['status']['name']
        status_at_sprint_end = status
        changelog = issue.get('changelog', {}).get('histories', [])
        
        if sprint_end_dt:
            # Ordenar el changelog por fecha para procesar cronológicamente
            sorted_changelog = sorted(changelog, key=lambda x: x.get('created', ''))
            
            # Buscar el estado inicial y el último estado antes del cierre del sprint
            initial_status = None
            last_status_before_sprint_end = None
            
            for history in sorted_changelog:
                change_date = history.get('created')
                if change_date:
                    try:
                        change_dt = datetime.strptime(change_date.split('.')[0], '%Y-%m-%dT%H:%M:%S')
                        
                        for item in history.get('items', []):
                            if item.get('field') == 'status':
                                # Guardar el estado inicial (fromString del primer cambio)
                                if initial_status is None:
                                    initial_status = item.get('fromString')
                                
                                # Si este cambio ocurrió antes o durante el sprint, actualizar el último estado
                                if change_dt <= sprint_end_dt:
                                    last_status_before_sprint_end = item.get('toString')
                        
                    except Exception:
                        pass
            
            # Determinar el estado al cierre del sprint
            if last_status_before_sprint_end:
                # Si hubo cambios antes del cierre, usar el último estado
                status_at_sprint_end = last_status_before_sprint_end
            elif initial_status:
                # Si no hubo cambios antes del cierre, usar el estado inicial
                status_at_sprint_end = initial_status
            else:
                # Si no hay changelog, usar el estado actual como fallback
                status_at_sprint_end = status
        
        time_spent = sum(worklog['timeSpentHours'] for worklog in issue['worklogs'])
        
        # Manejo explícito de Story Points usando el campo correcto
        story_points = issue['fields'].get('customfield_10030', 0)
        if story_points is not None:
            story_points = float(story_points)
        else:
            story_points = 0.0
            
        # Solo analizar story points vs tiempo si la issue está finalizada (Code Review, For Release, Done)
        if status_at_sprint_end in ['CODE REVIEW', 'For Release', 'Done']:
            story_points_analysis = analyze_story_points_vs_time(story_points, time_spent)
        else:
            story_points_analysis = ""
        
        parent_summary = ""
        if issue['fields'].get('parent'):
            parent_fields = issue['fields']['parent'].get('fields', {})
            parent_summary = parent_fields.get('summary', '')
        
        # Obtener la fecha de creación
        fecha_creacion = issue['fields'].get('created', '')
        if fecha_creacion:
            try:
                # Convertir a formato legible
                fecha_dt = datetime.strptime(fecha_creacion.split('.')[0], '%Y-%m-%dT%H:%M:%S')
                fecha_creacion = fecha_dt.strftime('%Y-%m-%d %H:%M:%S')
            except:
                fecha_creacion = fecha_creacion
        
        row = [
            issue_type,
            issue_key,
            summary,
            assignee,
            status_at_sprint_end,
            time_spent,  # valor numérico - worklogs ya filtrados por sprint
            story_points if story_points > 0 else None,  # valor numérico o celda vacía
            story_points_analysis,
            parent_summary,
            fecha_creacion
        ]
        ws.append(row)

    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f"{sprint_name}_sprint_analysis.xlsx"
    )

@app.route('/api/sprints/<int:sprint_id>/analysis/download_csv')
def download_sprint_analysis_csv(sprint_id):
    issues = get_issues_with_details(sprint_id)
    sprint_details = get_sprint_details(sprint_id)
    sprint_name = get_sprint_name(sprint_id)
    
    headers = [
        "Issue Type",
        "Issue Key",
        "Summary",
        "Assignee",
        "Status",
        "Time Spent",
        "Story Points",
        "Story Points vs Time",
        "Parent Summary",
        "Fecha de Creación"
    ]
    
    # Obtener fecha de fin del sprint como datetime
    sprint_end = sprint_details.get('endDate')
    if sprint_end:
        sprint_end_date = datetime.strptime(sprint_end, '%Y-%m-%d %H:%M:%S').date()
        sprint_end_dt = datetime.combine(sprint_end_date, datetime.max.time().replace(hour=23, minute=59, second=59, microsecond=0))
    else:
        sprint_end_dt = None
    
    def generate():
        output = []
        output.append(headers)
        for issue in issues:
            issue_type = issue['fields']['issuetype']['name']
            issue_key = issue['key']
            summary = issue['fields'].get('summary', '')
            assignee = ''
            if issue['fields'].get('assignee'):
                assignee = issue['fields']['assignee'].get('displayName', '')
            # Estado al final del sprint usando changelog
            status = issue['fields']['status']['name']
            status_at_sprint_end = status
            changelog = issue.get('changelog', {}).get('histories', [])
            
            if sprint_end_dt:
                # Ordenar el changelog por fecha para procesar cronológicamente
                sorted_changelog = sorted(changelog, key=lambda x: x.get('created', ''))
                
                # Buscar el estado inicial y el último estado antes del cierre del sprint
                initial_status = None
                last_status_before_sprint_end = None
                
                for history in sorted_changelog:
                    change_date = history.get('created')
                    if change_date:
                        try:
                            change_dt = datetime.strptime(change_date.split('.')[0], '%Y-%m-%dT%H:%M:%S')
                            
                            for item in history.get('items', []):
                                if item.get('field') == 'status':
                                    # Guardar el estado inicial (fromString del primer cambio)
                                    if initial_status is None:
                                        initial_status = item.get('fromString')
                                    
                                    # Si este cambio ocurrió antes o durante el sprint, actualizar el último estado
                                    if change_dt <= sprint_end_dt:
                                        last_status_before_sprint_end = item.get('toString')
                            
                        except Exception:
                            pass
                
                # Determinar el estado al cierre del sprint
                if last_status_before_sprint_end:
                    # Si hubo cambios antes del cierre, usar el último estado
                    status_at_sprint_end = last_status_before_sprint_end
                elif initial_status:
                    # Si no hubo cambios antes del cierre, usar el estado inicial
                    status_at_sprint_end = initial_status
                else:
                    # Si no hay changelog, usar el estado actual como fallback
                    status_at_sprint_end = status
            # Los worklogs ya están filtrados por sprint en get_issues_with_details
            time_spent = sum(worklog['timeSpentHours'] for worklog in issue['worklogs'])
            story_points = issue['fields'].get('customfield_10030', 0)
            if story_points is not None:
                story_points = float(story_points)
            else:
                story_points = 0.0
            # Solo analizar story points vs tiempo si la issue está finalizada (Code Review, For Release, Done)
            if status_at_sprint_end in ['CODE REVIEW', 'For Release', 'Done']:
                story_points_analysis = analyze_story_points_vs_time(story_points, time_spent)
            else:
                story_points_analysis = ""
            parent_summary = ""
            if issue['fields'].get('parent'):
                parent_fields = issue['fields']['parent'].get('fields', {})
                parent_summary = parent_fields.get('summary', '')
            # Obtener la fecha de creación
            fecha_creacion = issue['fields'].get('created', '')
            if fecha_creacion:
                try:
                    # Convertir a formato legible
                    fecha_dt = datetime.strptime(fecha_creacion.split('.')[0], '%Y-%m-%dT%H:%M:%S')
                    fecha_creacion = fecha_dt.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    fecha_creacion = fecha_creacion
            row = [
                issue_type,
                issue_key,
                summary,
                assignee,
                status_at_sprint_end,
                time_spent,
                story_points if story_points > 0 else None,
                story_points_analysis,
                parent_summary,
                fecha_creacion
            ]
            output.append(row)
        # Escribir CSV en memoria
        si = io.StringIO()
        writer = csv.writer(si)
        for row in output:
            writer.writerow(row)
        return si.getvalue()
    csv_data = generate()
    return Response(
        csv_data,
        mimetype='text/csv',
        headers={
            'Content-Disposition': f'attachment;filename={sprint_name}_sprint_analysis.csv'
        }
    )

@app.route('/metrics')
def metrics_view():
    return render_template('metrics.html')

@app.route('/comparative-metrics')
def comparative_metrics_view():
    return render_template('comparative_metrics.html')

@app.route('/api/sprints/<int:sprint_id>/advanced-metrics')
def get_advanced_metrics(sprint_id):
    try:
        issues = get_issues_with_details(sprint_id)
        sprint_details = get_sprint_details(sprint_id)
        
        # Mejorar el cálculo de la velocidad
        velocity_data = calculate_velocity_metrics(issues)
        
        metrics = {
            'velocity': velocity_data,
            'efficiency': {
                'stories_completed': velocity_data['completed_stories'],
                'total_stories': velocity_data['total_stories']
            },
            'time_analysis': {
                'time_distribution': calculate_time_distribution(issues),
                'average_task_completion': calculate_average_task_completion(issues)
            },
            'team_performance': calculate_team_performance(issues),
            'scope_changes': {
                'added': 0,  # Estos valores se pueden calcular si tienes los campos necesarios en JIRA
                'removed': 0
            }
        }
        
        return jsonify(metrics)
    except Exception as e:
        print(f"Error calculating metrics: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/metrics/velocity/<int:board_id>')
def get_velocity_metrics(board_id):
    """Obtener métricas de velocidad de los últimos 5 sprints completados"""
    try:
        sprints = get_sprints_for_board(board_id)
        # Filtrar solo sprints cerrados/completados
        completed_sprints = [s for s in sprints if s['state'].upper() in ['CLOSED']]
        recent_sprints = sorted(completed_sprints, key=lambda x: x['startDate'] if x.get('startDate') else '', reverse=True)[:5]
        
        velocity_data = {
            'sprints': [],
            'completed_points': [],
            'average': 0
        }
        
        total_points = 0
        for sprint in recent_sprints:
            issues = get_issues_with_details(sprint['id'])
            completed_points = sum(
                float(issue['fields'].get('customfield_10030', 0) or 0)
                for issue in issues
                if issue['fields']['status']['name'].upper() in ['DONE', 'CLOSED', 'FOR RELEASE']
            )
            
            velocity_data['sprints'].append(sprint['name'])
            velocity_data['completed_points'].append(completed_points)
            total_points += completed_points
        
        # Calcular el promedio
        velocity_data['average'] = total_points / len(recent_sprints) if recent_sprints else 0
        
        return jsonify(velocity_data)
    except Exception as e:
        print(f"Error calculating velocity metrics: {str(e)}")
        return jsonify({'error': str(e)}), 500

def calculate_velocity_metrics(issues):
    """
    Calcula métricas de velocidad detalladas del sprint
    """
    velocity_data = {
        'committed_points': 0,    # Puntos comprometidos al inicio
        'completed_points': 0,    # Puntos completados
        'total_stories': 0,
        'completed_stories': 0,
        'story_details': []       # Para debugging
    }

    for issue in issues:
        # Solo considerar historias y tareas técnicas (excluyendo Support, Bug y Spike que no se estiman en puntos)
        if issue['fields']['issuetype']['name'] in ['Story', 'Task']:
            # Obtener puntos comprometidos y completados
            story_points = issue['fields'].get('customfield_10030') or issue['fields'].get('customfield_10016', 0)
            
            if story_points:
                story_points = float(story_points)
                velocity_data['committed_points'] += story_points
                velocity_data['total_stories'] += 1
                
                story_detail = {
                    'key': issue['key'],
                    'summary': issue['fields']['summary'],
                    'type': issue['fields']['issuetype']['name'],
                    'status': issue['fields']['status']['name'],
                    'committed_points': story_points,
                    'completed_points': story_points if issue['fields']['status']['name'].upper() in ['DONE', 'CLOSED', 'FOR RELEASE'] else 0
                }
                
                if issue['fields']['status']['name'].upper() in ['DONE', 'CLOSED', 'FOR RELEASE']:
                    velocity_data['completed_points'] += story_points
                    velocity_data['completed_stories'] += 1
                    story_detail['completed'] = True
                else:
                    story_detail['completed'] = False
                
                velocity_data['story_details'].append(story_detail)

    return velocity_data

def calculate_time_distribution(issues):
    status_mapping = {
        'TO DO': 'planning',
        'TODO': 'planning',
        'IN PROGRESS': 'development',
        'CODE REVIEW': 'completed',
        'FOR RELEASE': 'completed',
        'DONE': 'completed',
    }
    
    distribution = {
        'planning': 0,
        'development': 0,
        'review': 0,
        'testing': 0,
        'completed': 0,
        'other': 0  # Agregamos 'other' como categoría válida
    }
    
    for issue in issues:
        current_status = issue['fields']['status']['name'].upper()
        mapped_status = status_mapping.get(current_status, 'other')
        
        for worklog in issue.get('worklogs', []):
            distribution[mapped_status] += worklog['timeSpentHours']
    
    # Solo incluir categorías con horas > 0
    return {k: v for k, v in distribution.items() if v > 0}

def calculate_average_task_completion(issues):
    completed_issues = [i for i in issues if i['fields']['status']['name'] in ['Done', 'Closed']]
    if not completed_issues:
        return 0
    
    total_days = 0
    count = 0
    
    for issue in completed_issues:
        if 'created' in issue['fields'] and 'resolutiondate' in issue['fields']:
            try:
                created = datetime.strptime(issue['fields']['created'].split('.')[0], '%Y-%m-%dT%H:%M:%S')
                resolved = datetime.strptime(issue['fields']['resolutiondate'].split('.')[0], '%Y-%m-%dT%H:%M:%S')
                days = (resolved - created).days
                if days >= 0:  # Ignorar valores negativos
                    total_days += days
                    count += 1
            except (ValueError, AttributeError):
                continue
    
    return total_days / count if count > 0 else 0

def calculate_team_performance(issues):
    team_performance = {}
    for issue in issues:
        for worklog in issue.get('worklogs', []):
            author = worklog['author']['displayName']
            if author not in team_performance:
                team_performance[author] = {'total_hours': 0, 'tasks_count': 0}
            team_performance[author]['total_hours'] += worklog['timeSpentHours']
            team_performance[author]['tasks_count'] += 1

    # Calcular promedio de horas por tarea para cada miembro
    for member in team_performance:
        stats = team_performance[member]
        stats['avg_hours_per_task'] = stats['total_hours'] / stats['tasks_count'] if stats['tasks_count'] > 0 else 0

    return team_performance

@app.route('/active')
def active_sprint_view():
    return render_template('active_sprint.html')

@app.route('/api/sprints/active/<int:board_id>')
def get_active_sprint(board_id):
    try:
        sprints = get_sprints_for_board(board_id)
        active_sprint = next((s for s in sprints if s['state'].upper() == 'ACTIVE'), None)
        
        if not active_sprint:
            return jsonify({'error': 'No active sprint found'}), 404
            
        issues = get_issues_with_details(active_sprint['id'])
        
        # Calcular métricas del sprint activo
        metrics = {
            'sprint': active_sprint,
            'total_issues': len(issues),
            'issues_by_type': {},
            'issues_by_status': {},
            'story_points': {
                'total': 0,
                'completed': 0
            },
            'burndown_data': calculate_burndown_data(active_sprint, issues)
        }
        
        # Contar issues por tipo
        for issue in issues:
            issue_type = issue['fields']['issuetype']['name']
            metrics['issues_by_type'][issue_type] = metrics['issues_by_type'].get(issue_type, 0) + 1
            
            # Contar por estado
            status = issue['fields']['status']['name']
            metrics['issues_by_status'][status] = metrics['issues_by_status'].get(status, 0) + 1
            
            # Sumar story points
            story_points = float(issue['fields'].get('customfield_10030', 0) or 0)
            metrics['story_points']['total'] += story_points
            if status.upper() in ['DONE', 'CLOSED', 'FOR RELEASE']:
                metrics['story_points']['completed'] += story_points
        
        return jsonify(metrics)
    except Exception as e:
        print(f"Error getting active sprint: {str(e)}")
        return jsonify({'error': str(e)}), 500

def calculate_burndown_data(sprint, issues):
    """Calcula los datos para el burndown chart considerando issues completados"""
    from datetime import datetime, timedelta
    
    COMPLETED_STATUSES = ['DONE', 'FOR RELEASE']
    
    start_date = datetime.strptime(sprint['startDate'].split('T')[0], '%Y-%m-%d')
    end_date = datetime.strptime(sprint['endDate'].split('T')[0], '%Y-%m-%d')
    total_days = (end_date - start_date).days + 1
    today = datetime.now().date()
    
    # Calcular total de story points y puntos completados
    total_points = 0
    completed_points = 0
    
    # Primero calculamos los totales
    for issue in issues:
        story_points = float(issue['fields'].get('customfield_10030', 0) or 0)
        if story_points > 0:
            total_points += story_points
            if issue['fields']['status']['name'].upper() in COMPLETED_STATUSES:
                completed_points += story_points

    # Los puntos restantes son los totales menos los completados
    remaining_points = total_points - completed_points
    
    ideal_burn = []
    actual_burn = []
    
    # Calcular línea ideal
    points_per_day = total_points / total_days
    for day in range(total_days):
        date = start_date + timedelta(days=day)
        ideal_burn.append({
            'date': date.strftime('%Y-%m-%d'),
            'points': round(total_points - (points_per_day * day), 1)
        })
    
    # Calcular línea actual
    current_date = start_date.date()
    current_points = total_points
    
    while current_date <= min(today, end_date.date()):
        # Para cada día, revisamos las issues completadas en esa fecha
        points_completed_today = 0
        for issue in issues:
            if (issue['fields']['status']['name'].upper() in COMPLETED_STATUSES and
                issue['fields'].get('resolutiondate')):
                resolution_date = datetime.strptime(
                    issue['fields']['resolutiondate'].split('T')[0], 
                    '%Y-%m-%d'
                ).date()
                if resolution_date == current_date:
                    story_points = float(issue['fields'].get('customfield_10030', 0) or 0)
                    points_completed_today += story_points
        
        current_points -= points_completed_today
        
        actual_burn.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'points': round(current_points, 1)
        })
        current_date += timedelta(days=1)
    
    # Debug info
    print(f"Total points: {total_points}")
    print(f"Completed points: {completed_points}")
    print(f"Remaining points: {remaining_points}")
    
    return {
        'ideal': ideal_burn,
        'actual': actual_burn,
        'total_points': total_points,
        'remaining_points': remaining_points
    }

@app.route('/api/metrics/summary/<int:board_id>')
def get_sprint_metrics_summary(board_id):
    """
    Devuelve un resumen de métricas para los últimos 5 sprints cerrados de un tablero:
    - Nombre y fechas del sprint
    - Story points comprometidos y completados
    - Ratio de fiabilidad (Say/Do)
    - Distribución de issues por tipo
    - Métricas de bugs (creados, resueltos, severidad, tiempo de resolución)
    """
    try:
        sprints = get_sprints_for_board(board_id)
        # Filtrar solo sprints cerrados/completados
        completed_sprints = [s for s in sprints if s['state'].upper() in ['CLOSED']]
        recent_sprints = sorted(completed_sprints, key=lambda x: x['startDate'] if x.get('startDate') else '', reverse=True)[:5]
        summary = []
        for sprint in recent_sprints:
            sprint_id = sprint['id']
            sprint_name = sprint['name']
            start_date = sprint.get('startDate')
            end_date = sprint.get('endDate')
            issues = get_issues_with_details(sprint_id)
            # Story points comprometidos y completados
            committed_points = 0
            completed_points = 0
            committed_issues = 0
            completed_issues = 0
            issue_type_dist = {}
            bugs_created = 0
            bugs_resolved = 0
            bug_severity = {}
            bug_resolution_times = []
            support_created = 0
            support_resolved = 0
            support_priority = {}
            support_resolution_times = []
            for issue in issues:
                print(f"Issue: {issue}")
                issue_type = issue['fields']['issuetype']['name']
                print(f"Issue Type: {issue['fields']['issuetype']}")
                issue_status = issue['fields']['status']['name']
                # Distribución de tipos
                if issue_type not in issue_type_dist:
                    issue_type_dist[issue_type] = 0
                issue_type_dist[issue_type] += 1
                # Story points
                story_points = issue['fields'].get('customfield_10030') or issue['fields'].get('customfield_10016', 0)
                if story_points:
                    try:
                        story_points = float(story_points)
                    except:
                        story_points = 0
                else:
                    story_points = 0
                # Comprometidos: todos los issues con story points (excluyendo Support, Bug y Spike que no se estiman en puntos)
                if issue_type in ['Story', 'Task']:
                    committed_points += story_points
                    committed_issues += 1
                    # Completados: solo los que están en estado final
                    if issue_status.upper() in ['DONE', 'CLOSED', 'FOR RELEASE']:
                        completed_points += story_points
                        completed_issues += 1
                # Bugs
                if issue_type == 'Bug':
                    bugs_created += 1
                    # Severidad
                    severity = issue['fields'].get('priority', {}).get('name', 'Sin severidad')
                    if severity not in bug_severity:
                        bug_severity[severity] = 0
                    bug_severity[severity] += 1
                    # Resueltos
                    if issue_status.upper() in ['DONE', 'CLOSED', 'FOR RELEASE']:
                        bugs_resolved += 1
                        # Tiempo de resolución
                        created = issue['fields'].get('created')
                        resolved = issue['fields'].get('resolutiondate')
                        if created and resolved:
                            try:
                                created_dt = datetime.strptime(created.split('.')[0], '%Y-%m-%dT%H:%M:%S')
                                resolved_dt = datetime.strptime(resolved.split('.')[0], '%Y-%m-%dT%H:%M:%S')
                                days = (resolved_dt - created_dt).days
                                if days >= 0:
                                    bug_resolution_times.append(days)
                            except Exception:
                                pass
                
                # Support
                if issue_type == 'Support':
                    support_created += 1
                    # Prioridad
                    priority = issue['fields'].get('priority', {}).get('name', 'Sin prioridad')
                    if priority not in support_priority:
                        support_priority[priority] = 0
                    support_priority[priority] += 1
                    # Resueltos
                    if issue_status.upper() in ['DONE', 'CLOSED', 'FOR RELEASE']:
                        support_resolved += 1
                        # Tiempo de resolución
                        created = issue['fields'].get('created')
                        resolved = issue['fields'].get('resolutiondate')
                        if created and resolved:
                            try:
                                created_dt = datetime.strptime(created.split('.')[0], '%Y-%m-%dT%H:%M:%S')
                                resolved_dt = datetime.strptime(resolved.split('.')[0], '%Y-%m-%dT%H:%M:%S')
                                days = (resolved_dt - created_dt).days
                                if days >= 0:
                                    support_resolution_times.append(days)
                            except Exception:
                                pass
            # Ratio de fiabilidad
            say_do_ratio = (completed_points / committed_points * 100) if committed_points > 0 else 0
            avg_bug_resolution = sum(bug_resolution_times) / len(bug_resolution_times) if bug_resolution_times else 0
            avg_support_resolution = sum(support_resolution_times) / len(support_resolution_times) if support_resolution_times else 0
            summary.append({
                'sprint_id': sprint_id,
                'sprint_name': sprint_name,
                'start_date': start_date,
                'end_date': end_date,
                'committed_points': committed_points,
                'completed_points': completed_points,
                'say_do_ratio': say_do_ratio,
                'issue_type_distribution': issue_type_dist,
                'bugs': {
                    'created': bugs_created,
                    'resolved': bugs_resolved,
                    'severity': bug_severity,
                    'avg_resolution_days': avg_bug_resolution
                },
                'support': {
                    'created': support_created,
                    'resolved': support_resolved,
                    'priority': support_priority,
                    'avg_resolution_days': avg_support_resolution
                }
            })
        return jsonify(summary)
    except Exception as e:
        print(f"Error in metrics summary: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/metrics/comparative', methods=['POST'])
def get_comparative_metrics():
    """
    Analiza múltiples sprints y devuelve métricas comparativas detalladas
    """
    try:
        sprint_ids = request.json.get('sprint_ids', [])
        if not sprint_ids or len(sprint_ids) > 10:
            return jsonify({'error': 'Se requieren entre 1 y 10 sprints'}), 400
        
        sprints_data = []
        
        for sprint_id in sprint_ids:
            sprint_details = get_sprint_details(sprint_id)
            issues = get_issues_with_details(sprint_id)
            
            # Calcular métricas del sprint
            sprint_metrics = calculate_comprehensive_sprint_metrics(sprint_details, issues)
            sprints_data.append(sprint_metrics)
        
        return jsonify({
            'sprints': sprints_data,
            'analysis_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        print(f"Error in comparative metrics: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/metrics/comparative/download_xlsx')
def download_comparative_analysis_xlsx():
    """
    Descarga análisis comparativo en formato Excel
    """
    try:
        sprint_ids = request.args.get('sprint_ids', '').split(',')
        if not sprint_ids or sprint_ids[0] == '':
            return jsonify({'error': 'Se requieren IDs de sprints'}), 400
        
        sprint_ids = [int(sid) for sid in sprint_ids]
        
        wb = Workbook()
        
        # Hoja de Resumen Ejecutivo
        ws_summary = wb.active
        ws_summary.title = "Resumen Ejecutivo"
        ws_summary.append(["Métricas Comparativas - Resumen Ejecutivo"])
        ws_summary.append([""])
        
        # Obtener datos de todos los sprints
        sprints_data = []
        for sprint_id in sprint_ids:
            sprint_details = get_sprint_details(sprint_id)
            issues = get_issues_with_details(sprint_id)
            sprint_metrics = calculate_comprehensive_sprint_metrics(sprint_details, issues)
            sprints_data.append(sprint_metrics)
        
        # Generar insights
        insights = generate_executive_insights(sprints_data)
        ws_summary.append(["Insights Clave:"])
        for insight in insights:
            ws_summary.append([insight])
        
        ws_summary.append([""])
        ws_summary.append(["Métricas por Sprint:"])
        ws_summary.append(["Sprint", "Story Points Completados", "Horas Totales", "Eficiencia (SP/Hora)", "Ratio Bugs/Features", "Ratio Support/Features", "Precisión Estimaciones"])
        
        for sprint in sprints_data:
            efficiency = sprint['total_hours'] > 0 and (sprint['completed_points'] / sprint['total_hours']) or 0
            bugs_ratio = calculate_bugs_ratio(sprint)
            support_ratio = calculate_support_ratio(sprint)
            estimation_accuracy = calculate_estimation_accuracy(sprint)
            
            ws_summary.append([
                sprint['name'],
                sprint['completed_points'],
                round(sprint['total_hours'], 1),
                round(efficiency, 2),
                f"{bugs_ratio:.1f}%",
                f"{support_ratio:.1f}%",
                f"{estimation_accuracy:.1f}%"
            ])
        
        # Hoja de Métricas Individuales
        ws_individual = wb.create_sheet(title="Métricas Individuales")
        ws_individual.append(["Desarrollador", "Story Points Completados", "Horas Trabajadas", "Eficiencia (SP/Hora)", "Tareas Completadas"])
        
        individual_data = aggregate_individual_metrics(sprints_data)
        for developer, metrics in individual_data.items():
            efficiency = metrics['total_hours'] > 0 and (metrics['completed_points'] / metrics['total_hours']) or 0
            ws_individual.append([
                developer,
                metrics['completed_points'],
                round(metrics['total_hours'], 1),
                round(efficiency, 2),
                metrics['completed_tasks']
            ])
        
        # Hoja de Análisis Detallado
        ws_detailed = wb.create_sheet(title="Análisis Detallado")
        ws_detailed.append([
            "Sprint", "Issue Type", "Issue Key", "Summary", "Assignee", "Status", 
            "Time Spent", "Story Points", "Story Points vs Time", "Parent Summary", "Fecha de Creación"
        ])
        
        for sprint in sprints_data:
            for issue in sprint['detailed_issues']:
                ws_detailed.append([
                    sprint['name'],
                    issue['issue_type'],
                    issue['issue_key'],
                    issue['summary'],
                    issue['assignee'],
                    issue['status'],
                    issue['time_spent'],
                    issue['story_points'] if issue['story_points'] > 0 else None,
                    issue['story_points_analysis'],
                    issue['parent_summary'],
                    issue['fecha_creacion']
                ])
        
        # Ajustar ancho de columnas
        for ws in [ws_summary, ws_individual, ws_detailed]:
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        return send_file(
            file_stream,
            as_attachment=True,
            download_name=f"comparative_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
    except Exception as e:
        print(f"Error downloading comparative analysis: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/metrics/comparative/download_csv')
def download_comparative_analysis_csv():
    """
    Descarga análisis comparativo en formato CSV
    """
    try:
        sprint_ids = request.args.get('sprint_ids', '').split(',')
        if not sprint_ids or sprint_ids[0] == '':
            return jsonify({'error': 'Se requieren IDs de sprints'}), 400
        
        sprint_ids = [int(sid) for sid in sprint_ids]
        
        # Obtener datos de todos los sprints
        sprints_data = []
        for sprint_id in sprint_ids:
            sprint_details = get_sprint_details(sprint_id)
            issues = get_issues_with_details(sprint_id)
            sprint_metrics = calculate_comprehensive_sprint_metrics(sprint_details, issues)
            sprints_data.append(sprint_metrics)
        
        def generate_csv():
            output = []
            
            # Resumen ejecutivo
            output.append(["Métricas Comparativas - Resumen Ejecutivo"])
            output.append([""])
            
            insights = generate_executive_insights(sprints_data)
            output.append(["Insights Clave:"])
            for insight in insights:
                output.append([insight])
            
            output.append([""])
            output.append(["Métricas por Sprint:"])
            output.append(["Sprint", "Story Points Completados", "Horas Totales", "Eficiencia (SP/Hora)", "Ratio Bugs/Features", "Ratio Support/Features", "Precisión Estimaciones"])
            
            for sprint in sprints_data:
                efficiency = sprint['total_hours'] > 0 and (sprint['completed_points'] / sprint['total_hours']) or 0
                bugs_ratio = calculate_bugs_ratio(sprint)
                support_ratio = calculate_support_ratio(sprint)
                estimation_accuracy = calculate_estimation_accuracy(sprint)
                
                output.append([
                    sprint['name'],
                    sprint['completed_points'],
                    round(sprint['total_hours'], 1),
                    round(efficiency, 2),
                    f"{bugs_ratio:.1f}%",
                    f"{support_ratio:.1f}%",
                    f"{estimation_accuracy:.1f}%"
                ])
            
            output.append([""])
            output.append(["Métricas Individuales:"])
            output.append(["Desarrollador", "Story Points Completados", "Horas Trabajadas", "Eficiencia (SP/Hora)", "Tareas Completadas"])
            
            individual_data = aggregate_individual_metrics(sprints_data)
            for developer, metrics in individual_data.items():
                efficiency = metrics['total_hours'] > 0 and (metrics['completed_points'] / metrics['total_hours']) or 0
                output.append([
                    developer,
                    metrics['completed_points'],
                    round(metrics['total_hours'], 1),
                    round(efficiency, 2),
                    metrics['completed_tasks']
                ])
            
            output.append([""])
            output.append(["Análisis Detallado:"])
            output.append([
                "Sprint", "Issue Type", "Issue Key", "Summary", "Assignee", "Status", 
                "Time Spent", "Story Points", "Story Points vs Time", "Parent Summary", "Fecha de Creación"
            ])
            
            for sprint in sprints_data:
                for issue in sprint['detailed_issues']:
                    output.append([
                        sprint['name'],
                        issue['issue_type'],
                        issue['issue_key'],
                        issue['summary'],
                        issue['assignee'],
                        issue['status'],
                        issue['time_spent'],
                        issue['story_points'] if issue['story_points'] > 0 else None,
                        issue['story_points_analysis'],
                        issue['parent_summary'],
                        issue['fecha_creacion']
                    ])
            
            # Escribir CSV en memoria
            si = io.StringIO()
            writer = csv.writer(si)
            for row in output:
                writer.writerow(row)
            return si.getvalue()
        
        csv_data = generate_csv()
        return Response(
            csv_data,
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment;filename=comparative_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            }
        )
        
    except Exception as e:
        print(f"Error downloading comparative analysis CSV: {str(e)}")
        return jsonify({'error': str(e)}), 500

def calculate_comprehensive_sprint_metrics(sprint_details, issues):
    """
    Calcula métricas comprehensivas para un sprint
    """
    # Obtener fecha de fin del sprint
    sprint_end = sprint_details.get('endDate')
    sprint_end_dt = None
    if sprint_end:
        sprint_end_date = datetime.strptime(sprint_end, '%Y-%m-%d %H:%M:%S').date()
        sprint_end_dt = datetime.combine(sprint_end_date, datetime.max.time().replace(hour=23, minute=59, second=59, microsecond=0))
    
    # Métricas básicas
    completed_points = 0
    estimated_points = 0  # Puntos estimados totales del sprint
    total_hours = 0
    issue_type_distribution = {}
    individual_metrics = {}
    detailed_issues = []
    estimation_analysis = []
    
    for issue in issues:
        issue_type = issue['fields']['issuetype']['name']
        issue_key = issue['key']
        summary = issue['fields'].get('summary', '')
        
        # Assignee
        assignee = ''
        if issue['fields'].get('assignee'):
            assignee = issue['fields']['assignee'].get('displayName', '')
        
        # Estado al final del sprint usando changelog
        status = issue['fields']['status']['name']
        status_at_sprint_end = status
        changelog = issue.get('changelog', {}).get('histories', [])
        
        if sprint_end_dt:
            sorted_changelog = sorted(changelog, key=lambda x: x.get('created', ''))
            initial_status = None
            last_status_before_sprint_end = None
            
            for history in sorted_changelog:
                change_date = history.get('created')
                if change_date:
                    try:
                        change_dt = datetime.strptime(change_date.split('.')[0], '%Y-%m-%dT%H:%M:%S')
                        
                        for item in history.get('items', []):
                            if item.get('field') == 'status':
                                if initial_status is None:
                                    initial_status = item.get('fromString')
                                
                                if change_dt <= sprint_end_dt:
                                    last_status_before_sprint_end = item.get('toString')
                    except Exception:
                        pass
            
            if last_status_before_sprint_end:
                status_at_sprint_end = last_status_before_sprint_end
            elif initial_status:
                status_at_sprint_end = initial_status
            else:
                status_at_sprint_end = status
        
        # Time spent - worklogs ya están filtrados por sprint en get_issues_with_details
        time_spent = sum(worklog['timeSpentHours'] for worklog in issue['worklogs'])
        total_hours += time_spent
        
        # Story points
        story_points = issue['fields'].get('customfield_10030', 0)
        if story_points is not None:
            story_points = float(story_points)
        else:
            story_points = 0.0
        
        # Contar puntos estimados totales (solo para Task y Story, excluyendo Support, Bug y Spike)
        if issue_type in ['Task', 'Story'] and story_points > 0:
            estimated_points += story_points
        
        # Solo contar story points completados si está en estado final
        if status_at_sprint_end in ['Done', 'For Release', 'CODE REVIEW']:
            completed_points += story_points
        
        # Distribución por tipo
        if issue_type not in issue_type_distribution:
            issue_type_distribution[issue_type] = 0
        issue_type_distribution[issue_type] += 1
        
        # Métricas individuales
        if assignee:
            if assignee not in individual_metrics:
                individual_metrics[assignee] = {
                    'developer': assignee,
                    'completed_points': 0,
                    'total_hours': 0,
                    'completed_tasks': 0
                }
            
            individual_metrics[assignee]['total_hours'] += time_spent
            if status_at_sprint_end in ['Done', 'For Release', 'CODE REVIEW']:
                individual_metrics[assignee]['completed_points'] += story_points
                individual_metrics[assignee]['completed_tasks'] += 1
        
        # Análisis de estimaciones
        if status_at_sprint_end in ['CODE REVIEW', 'For Release', 'Done']:
            analysis = analyze_story_points_vs_time(story_points, time_spent)
            estimation_analysis.append({
                'issue_key': issue_key,
                'story_points': story_points,
                'time_spent': time_spent,
                'analysis': analysis
            })
        
        # Parent summary
        parent_summary = ""
        if issue['fields'].get('parent'):
            parent_fields = issue['fields']['parent'].get('fields', {})
            parent_summary = parent_fields.get('summary', '')
        
        # Fecha de creación
        fecha_creacion = issue['fields'].get('created', '')
        if fecha_creacion:
            try:
                fecha_dt = datetime.strptime(fecha_creacion.split('.')[0], '%Y-%m-%dT%H:%M:%S')
                fecha_creacion = fecha_dt.strftime('%Y-%m-%d %H:%M:%S')
            except:
                pass
        
        # Detalles de la issue para el reporte
        detailed_issues.append({
            'issue_type': issue_type,
            'issue_key': issue_key,
            'summary': summary,
            'assignee': assignee,
            'status': status_at_sprint_end,
            'time_spent': time_spent,
            'story_points': story_points,
            'story_points_analysis': analyze_story_points_vs_time(story_points, time_spent) if status_at_sprint_end in ['CODE REVIEW', 'For Release', 'Done'] else "",
            'parent_summary': parent_summary,
            'fecha_creacion': fecha_creacion
        })
    
    return {
        'id': sprint_details['id'],
        'name': sprint_details['name'],
        'start_date': sprint_details.get('startDate'),
        'end_date': sprint_details.get('endDate'),
        'completed_points': completed_points,
        'estimated_points': estimated_points,  # Puntos estimados totales
        'total_hours': total_hours,
        'issue_type_distribution': issue_type_distribution,
        'individual_metrics': list(individual_metrics.values()),
        'detailed_issues': detailed_issues,
        'estimation_analysis': estimation_analysis
    }

def generate_executive_insights(sprints_data):
    """
    Genera insights ejecutivos basados en los datos de los sprints
    """
    insights = []
    
    if not sprints_data:
        return ["No hay datos suficientes para generar insights"]
    
    # Velocity insights
    velocities = [s['completed_points'] for s in sprints_data]
    avg_velocity = sum(velocities) / len(velocities)
    latest_velocity = velocities[-1]
    
    if latest_velocity < avg_velocity * 0.8:
        insights.append(f"⚠️ Velocity baja: El último sprint ({latest_velocity} SP) está {((avg_velocity - latest_velocity) / avg_velocity * 100):.1f}% por debajo del promedio")
    elif latest_velocity > avg_velocity * 1.2:
        insights.append(f"🚀 Velocity excelente: El último sprint superó el promedio en {((latest_velocity - avg_velocity) / avg_velocity * 100):.1f}%")
    
    # Bugs ratio insights
    latest_sprint = sprints_data[-1]
    bugs_ratio = calculate_bugs_ratio(latest_sprint)
    if bugs_ratio > 25:
        insights.append(f"🐛 Alto ratio de bugs: {bugs_ratio:.1f}% de las tareas son bugs")
    
    # Support ratio insights
    support_ratio = calculate_support_ratio(latest_sprint)
    if support_ratio > 30:
        insights.append(f"🆘 Alto ratio de support: {support_ratio:.1f}% de las tareas son support")
    
    # Estimation accuracy insights
    estimation_accuracy = calculate_estimation_accuracy(latest_sprint)
    if estimation_accuracy < 70:
        insights.append(f"📊 Baja precisión de estimaciones: {estimation_accuracy:.1f}% de las estimaciones fueron correctas")
    
    return insights

def calculate_bugs_ratio(sprint_data):
    """
    Calcula el ratio de bugs vs features
    """
    bugs = sprint_data['issue_type_distribution'].get('Bug', 0)
    features = sprint_data['issue_type_distribution'].get('Story', 0) + sprint_data['issue_type_distribution'].get('Task', 0)
    return (bugs / features * 100) if features > 0 else 0

def calculate_support_ratio(sprint_data):
    """
    Calcula el ratio de support vs features
    """
    support = sprint_data['issue_type_distribution'].get('Support', 0)
    features = sprint_data['issue_type_distribution'].get('Story', 0) + sprint_data['issue_type_distribution'].get('Task', 0)
    return (support / features * 100) if features > 0 else 0

def calculate_estimation_accuracy(sprint_data):
    """
    Calcula la precisión de las estimaciones
    """
    estimations = sprint_data['estimation_analysis']
    if not estimations:
        return 0
    
    correct = len([e for e in estimations if e['analysis'] == 'Correcto'])
    return (correct / len(estimations)) * 100

def aggregate_individual_metrics(sprints_data):
    """
    Agrega métricas individuales de todos los sprints
    """
    aggregated = {}
    
    for sprint in sprints_data:
        # individual_metrics es una lista de diccionarios con los datos
        # pero necesitamos reconstruir el mapeo de desarrollador -> métricas
        # Vamos a usar los detailed_issues para reconstruir esto
        developer_metrics = {}
        
        for issue in sprint['detailed_issues']:
            assignee = issue['assignee']
            if assignee:
                if assignee not in developer_metrics:
                    developer_metrics[assignee] = {
                        'completed_points': 0,
                        'total_hours': 0,
                        'completed_tasks': 0
                    }
                
                developer_metrics[assignee]['total_hours'] += issue['time_spent']
                
                if issue['status'] in ['Done', 'For Release', 'CODE REVIEW']:
                    developer_metrics[assignee]['completed_points'] += issue['story_points']
                    developer_metrics[assignee]['completed_tasks'] += 1
        
        # Agregar al total agregado
        for developer, metrics in developer_metrics.items():
            if developer not in aggregated:
                aggregated[developer] = {
                    'completed_points': 0,
                    'total_hours': 0,
                    'completed_tasks': 0
                }
            
            aggregated[developer]['completed_points'] += metrics['completed_points']
            aggregated[developer]['total_hours'] += metrics['total_hours']
            aggregated[developer]['completed_tasks'] += metrics['completed_tasks']
    
    return aggregated

if __name__ == '__main__':
    app.run(debug=True)
